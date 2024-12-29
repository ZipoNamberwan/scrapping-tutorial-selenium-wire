from seleniumwire import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import requests
from urllib.parse import urlparse, unquote, urlencode, urlunparse
import json
import openpyxl
from openpyxl.utils import get_column_letter
import os

options = webdriver.ChromeOptions()
# options.add_argument('--headless')
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)

while True:
   try:
    driver.get("http://127.0.0.1:8000/login")
    break
   except TimeoutException:
     driver.refresh()

wait_interval = 10
wait = WebDriverWait(driver, wait_interval)

while True:
    try:
        element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="email"]')))
        element.send_keys('user@gmail.com')
        element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="password"]')))
        element.send_keys('12345678')
        element = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div/div[2]/form/div[4]/button')))
        element.click()
        element = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="myTable_wrapper"]',)))
        time.sleep(5)
        break
    except TimeoutException:
        driver.refresh()

total_records = 0
records_per_page = 100

for request in driver.requests:
    if "http://127.0.0.1:8000/pegawai" in request.url:
        selected_request = request
        parsed_url = urlparse(selected_request.url)
        
        query_string = parsed_url.query
        query_params = {}
        for param in query_string.split('&'):
            key, value = param.split('=', 1) if '=' in param else (param, '')
            key = unquote(key)
            value = unquote(value)
            query_params[key] = value

        response_body = json.loads(request.response.body.decode("utf-8", errors="ignore"))

        total_records = response_body['recordsTotal']

def make_request(url, headers, payload=None, attempt=0, max_retries=10):
    try:
        r = requests.get(url, headers=headers, params=payload)
        return r
    except Exception:
        if attempt < max_retries:
            print(f"Request failed. Retrying...")
            return make_request(url, headers, attempt + 1, max_retries)
        else:
            print("Request failed after maximum number of retries.")
            return None

total_pages = total_records // records_per_page
if total_records % records_per_page != 0:
    total_pages += 1

employee = []
for page in range(total_pages):
    start = page * records_per_page
    query_params['start'] = [str(start)]
    query_params['length'] = [str(records_per_page)] 

    updated_query = urlencode(query_params, doseq=True)
    modified_url = urlunparse(parsed_url._replace(query=updated_query))

    response = make_request(modified_url, dict(selected_request.headers))
    data = response.json()['data']
    employee = employee + data
    print(page)

def save_data_to_excel(data_dict, file_path):
    wb = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    for sheet_name, data in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        # Check if data is not empty
        if not data:
            print(f"No data to write for sheet '{sheet_name}'.", flush=True)
            continue

        # Dynamically determine the column headers based on keys in the first dictionary
        headers = list(data[0].keys())

        # Write the headers to the first row
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        # Write the data rows
        for row_num, entry in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = entry.get(header, None)

    # Save the workbook to the specified file path
    wb.save(file_path)
    print(f"Data successfully saved to {file_path}", flush=True)

try:
    base_dir = os.path.dirname(os.path.abspath(__file__))
except NameError:
    base_dir = os.getcwd()

if "__file__" in globals():
    excel_path = os.path.join(base_dir)
else:
    excel_path = os.path.join(base_dir, "result")

data_dict = {
    "Data": employee,
}
save_data_to_excel(data_dict, excel_path + '\\' + "result.xlsx")